"""
ReplaceDisplayNameWithCNname.py

Description:
    This script processes an Excel file to replace displayName values in the 'Mapping' tab 
    with corresponding GroupName values from the 'non-Existing AD groups' tab.
    
    The script performs the following operations:
    1. Reads the 'non-Existing AD groups' tab and extracts displayName and GroupName columns
    2. Trims all values (removes leading/trailing whitespace) for accurate matching
    3. Creates a lookup mapping of displayName -> GroupName (for both empty and non-empty values)
    4. Searches through all cells in the 'Mapping' tab for matching displayName values (using trimmed comparison)
    5. When a match is found:
       - If GroupName is NOT empty: Replaces the cell value with GroupName (trimmed) and highlights in YELLOW
       - If GroupName IS empty (null, "", or " "): Highlights the cell in dark RED with YELLOW font
    6. Reorders sheets to: Mapping, no phase, non-Existing AD groups
    7. Saves the updated workbook with a new filename (*_updated.xlsx)
    
    Note: Empty GroupName values are now highlighted in dark RED with yellow font for visibility.
          All values are trimmed (whitespace removed) before comparison and insertion to ensure
          accurate matching and clean data quality.

Input:
    - Excel file: mapping_persona_sg/ID now-Personas mapping.xlsx
    - Required tabs: 'non-Existing AD groups', 'Mapping'
    - Required columns in 'non-Existing AD groups': displayName, GroupName

Output:
    - Updated Excel file saved as: *_updated.xlsx with replacements highlighted in yellow
    - Sheet order: 1. Mapping, 2. no phase, 3. non-Existing AD groups

Usage:
    python tools/ReplaceDisplayNameWithCNname.py
"""

import openpyxl
from openpyxl.styles import PatternFill, Font
import os

# ============================================================================
# CONFIGURATION
# ============================================================================

# Define relative path from project root to the Excel file
EXCEL_FILE = os.path.join("mapping_persona_sg", "ID now-Personas mapping.xlsx")

# Define yellow fill pattern for highlighting replaced cells
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Define dark red fill pattern for highlighting empty GroupName matches
DARK_RED_FILL = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")

# Define yellow font for text highlighting
YELLOW_FONT = Font(color="FFFF00")


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def is_empty(value):
    """
    Check if a value is considered empty.
    
    Args:
        value: Any value to check (can be None, string, or other type)
    
    Returns:
        bool: True if value is None, empty string, or whitespace-only string
    """
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


# ============================================================================
# MAIN PROCESSING FUNCTION
# ============================================================================

def main():
    """
    Main function to process the Excel file and perform replacements.
    
    Process flow:
    1. Load the Excel workbook
    2. Access required sheets ('non-Existing AD groups' and 'Mapping')
    3. Build a mapping dictionary from displayName to GroupName (with trimmed values)
    4. Iterate through all cells in 'Mapping' sheet (comparing trimmed values)
    5. Replace matching values with trimmed GroupName and highlight in yellow
    6. Reorder sheets: Mapping, no phase, non-Existing AD groups
    7. Save the updated workbook
    
    Note: All values are trimmed before comparison and insertion to ensure accurate matching.
    """
    
    # ========================================================================
    # STEP 1: Load the workbook
    # ========================================================================
    print(f"Loading workbook: {EXCEL_FILE}")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    
    # ========================================================================
    # STEP 2: Access the required sheets
    # ========================================================================
    try:
        non_existing_sheet = wb["non-Existing AD groups"]
        mapping_sheet = wb["Mapping"]
    except KeyError as e:
        print(f"Error: Sheet not found - {e}")
        print(f"Available sheets: {wb.sheetnames}")
        return
    
    # ========================================================================
    # STEP 3: Find column indices in 'non-Existing AD groups' sheet
    # ========================================================================
    # Read the header row (row 1) to find column positions
    non_existing_headers = [cell.value for cell in non_existing_sheet[1]]
    
    try:
        # Get column index for 'displayName' (add 1 because openpyxl uses 1-based indexing)
        display_name_col_idx = non_existing_headers.index("displayName") + 1
        # Get column index for 'GroupName'
        group_name_col_idx = non_existing_headers.index("GroupName") + 1
    except ValueError as e:
        print(f"Error: Required column not found in 'non-Existing AD groups' sheet - {e}")
        print(f"Available columns: {non_existing_headers}")
        return
    
    # ========================================================================
    # STEP 4: Build mapping dictionaries: displayName -> GroupName (trimmed)
    # ========================================================================
    # Dictionary for displayNames with valid GroupNames
    display_to_group_mapping = {}
    # Set for displayNames with empty GroupNames (to highlight in red)
    display_names_with_empty_groups = set()
    
    # Iterate through all data rows (starting from row 2, skipping header)
    for row_idx in range(2, non_existing_sheet.max_row + 1):
        display_name = non_existing_sheet.cell(row_idx, display_name_col_idx).value
        group_name = non_existing_sheet.cell(row_idx, group_name_col_idx).value
        
        # Trim displayName before processing
        if display_name is not None:
            display_name = str(display_name).strip()
        
        # Trim groupName before processing
        if group_name is not None:
            group_name = str(group_name).strip()
        
        # Skip if displayName itself is empty
        if is_empty(display_name):
            continue
            
        # Check if GroupName is empty
        if is_empty(group_name):
            # Track displayNames with empty GroupNames for red highlighting (trimmed value)
            display_names_with_empty_groups.add(display_name)
        else:
            # Add to mapping for replacement (both trimmed values)
            display_to_group_mapping[display_name] = group_name
    
    print(f"Found {len(display_to_group_mapping)} displayName -> GroupName mappings to process")
    print(f"Found {len(display_names_with_empty_groups)} displayNames with empty GroupNames to highlight in RED")
    
    # ========================================================================
    # STEP 5: Process the 'Mapping' sheet - search and replace (using trimmed values)
    # ========================================================================
    replacements_made = 0
    red_highlights_made = 0
    
    # Iterate through all cells in the 'Mapping' sheet
    for row in mapping_sheet.iter_rows(min_row=1, max_row=mapping_sheet.max_row):
        for cell in row:
            cell_value = cell.value
            
            # Trim cell value before comparison
            if cell_value is not None:
                cell_value_trimmed = str(cell_value).strip()
            else:
                cell_value_trimmed = None
            
            # Check if the trimmed cell value matches any displayName with a valid GroupName
            if cell_value_trimmed in display_to_group_mapping:
                group_name = display_to_group_mapping[cell_value_trimmed]
                
                # Replace the cell value with the corresponding GroupName (trimmed value)
                cell.value = group_name
                
                # Set the cell background to YELLOW for visual identification
                cell.fill = YELLOW_FILL
                
                # Track the number of replacements and log the change
                replacements_made += 1
                print(f"Replaced '{cell_value_trimmed}' with '{group_name}' at cell {cell.coordinate}")
            
            # Check if the trimmed cell value matches a displayName with empty GroupName
            elif cell_value_trimmed in display_names_with_empty_groups:
                # Set the cell background to dark RED and font to YELLOW
                cell.fill = DARK_RED_FILL
                cell.font = YELLOW_FONT
                
                # Track the number of red highlights and log the change
                red_highlights_made += 1
                print(f"Highlighted '{cell_value_trimmed}' in RED (empty GroupName) at cell {cell.coordinate}")
    
    print(f"\nTotal replacements made: {replacements_made}")
    print(f"Total red highlights (empty GroupName): {red_highlights_made}")
    
    # ========================================================================
    # STEP 6: Reorder sheets in the workbook
    # ========================================================================
    print(f"\nReordering sheets...")
    
    # Define desired sheet order
    desired_order = ["Mapping", "no phase", "non-Existing AD Groups"]
    
    # Get current sheets
    current_sheets = wb.sheetnames
    print(f"Current sheet order: {current_sheets}")
    
    # Reorder sheets according to desired order
    # Only reorder sheets that exist in the workbook
    ordered_sheets = []
    for sheet_name in desired_order:
        if sheet_name in current_sheets:
            ordered_sheets.append(sheet_name)
    
    # Add any remaining sheets that weren't in the desired order
    for sheet_name in current_sheets:
        if sheet_name not in ordered_sheets:
            ordered_sheets.append(sheet_name)
    
    # Reorder by moving sheets to their correct positions
    for i, sheet_name in enumerate(ordered_sheets):
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            wb.move_sheet(sheet, offset=i - wb.sheetnames.index(sheet_name))
    
    print(f"New sheet order: {wb.sheetnames}")
    
    # ========================================================================
    # STEP 7: Save the updated workbook
    # ========================================================================
    # Create output filename by appending '_updated' before the file extension
    output_file = EXCEL_FILE.replace(".xlsx", "_updated.xlsx")
    wb.save(output_file)
    print(f"\nWorkbook saved as: {output_file}")


# ============================================================================
# SCRIPT ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    # Execute main function when script is run directly
    main()
